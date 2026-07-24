import streamlit as st
import pandas as pd
import random
import datetime
import calendar
from collections import Counter
import io
from PIL import Image, ImageDraw, ImageFont

# Exportadores
try:
    import openpyxl
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    PDF_DISPONIBLE = True
except ImportError:
    PDF_DISPONIBLE = False

st.set_page_config(page_title="Programa Audio, Video y Salas", layout="centered", page_icon="📅")

st.title("📅 Generador Avanzado: Audio, Video y Salas")
st.caption("Congregación Gallito, San José de la Montaña")

if "programa_guardado" not in st.session_state:
    st.session_state["programa_guardado"] = None

# --- 1. Base de Hermanos por Rol ---
hermanos_av = [
    "Carlos Josué Pereira", "José Pereira", "Josué López", 
    "Rodney Alfaro", "Geremy Fernández", "Julio Sánchez", "Dashler Sánchez", 
    "Sebastián Montero", "David Herrera", "José Alberto González", "Javier García"
]

hermanos_solo_mics = [
    "Rafael Segura", "Kenneth Solís", "Walter Sánchez", 
    "Iván Zamora", "Carlos Blanco", "Elixander Alvarado"
]

ancianos_y_ministeriales = [
    "Carlos Josué Pereira", "Carlos Enrique Pereira", "José Pereira", "Josué López", "Rodney Alfaro",
    "Geremy Fernández", "Julio Sánchez", "David Herrera", "José Alberto González",
    "Javier García", "Elixander Alvarado", "Roger Loaiza", "Walter Sánchez"
]

FECHA_VISITA_SC = datetime.date(2026, 8, 23)

# --- 2. Opciones e Imágenes ---
st.sidebar.header("⚙️ Configuración del Programa")

# Carga opcional del Logo / Imagen
logo_subido = st.sidebar.file_uploader("🖼️ Cargar Logo/Imagen para Encabezado (Opcional)", type=["png", "jpg", "jpeg"])

anio = st.sidebar.number_input("Año", min_value=2026, max_value=2030, value=2026)
mes_nombres = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]
mes_sel = st.sidebar.selectbox("Seleccione el Mes de Inicio", mes_nombres, index=6)
num_mes = mes_nombres.index(mes_sel) + 1

duracion = st.sidebar.radio("¿Cuántos meses desea generar?", ["1 Mes", "2 Meses"], index=0)

# Cargar Historial
st.sidebar.divider()
st.sidebar.header("📂 Historial Anterior")
archivo_historial = st.sidebar.file_uploader("Cargar CSV del mes anterior (Opcional)", type=["csv"])

hermanos_recientes = []
if archivo_historial is not None:
    try:
        df_ant = pd.read_csv(archivo_historial)
        if not df_ant.empty:
            ultima_fila = df_ant.iloc[-1]
            for col in ["Audio", "Video", "Micrófono", "Acomodador"]:
                if col in df_ant.columns and "---" not in str(ultima_fila[col]):
                    hermanos_recientes.append(ultima_fila[col])
            st.sidebar.success("¡Historial anterior cargado con éxito!")
    except Exception:
        st.sidebar.warning("No se pudo leer el archivo de historial.")

def obtener_fechas_mes(m_num, a_num):
    n_dias = calendar.monthrange(a_num, m_num)[1]
    f_lista = []
    for dia in range(1, n_dias + 1):
        f_dt = datetime.date(a_num, m_num, dia)
        if f_dt.weekday() in [2, 6]: # Miércoles y Domingo
            f_lista.append(f_dt)
    return f_lista

fechas_reunion = obtener_fechas_mes(num_mes, anio)
if duracion == "2 Meses":
    sig_mes = num_mes + 1 if num_mes < 12 else 1
    sig_anio = anio if num_mes < 12 else anio + 1
    fechas_reunion += obtener_fechas_mes(sig_mes, sig_anio)

texto_rango = f"{mes_sel} {anio}" if duracion == "1 Mes" else f"{mes_sel} y {mes_nombres[(num_mes % 12)]} {anio}"
st.subheader(f"🗓️ Reuniones para {texto_rango}: {len(fechas_reunion)} fechas")

if logo_subido:
    st.image(logo_subido, width=140)

# --- 3. Formulario de Ocupaciones y Ajustes ---
ocupados_por_fecha = {}
canceladas_por_fecha = {}
dias_ajustados_por_fecha = {}

with st.expander("📌 Configurar fechas (Ocupados, Visita SC y Cancelaciones)", expanded=True):
    for f in fechas_reunion:
        es_miercoles = f.weekday() == 2
        dia_semana_defecto = "Miércoles" if es_miercoles else "Domingo"
        
        st.markdown(f"### 🗓️ {dia_semana_defecto} {f.strftime('%d/%m/%Y')}")
        
        if es_miercoles:
            dia_real = st.radio(
                f"Día de la reunión (Visita SC):",
                options=["Miércoles", "Martes"],
                index=0,
                key=f"dia_real_{f}",
                horizontal=True
            )
            fecha_efectiva = f - datetime.timedelta(days=1) if dia_real == "Martes" else f
        else:
            dia_real = "Domingo"
            fecha_efectiva = f
            
        dias_ajustados_por_fecha[f] = (dia_real, fecha_efectiva)
        
        es_cancelada = st.checkbox(f"❌ Cancelar esta reunión (Asamblea / Sin reunión)", key=f"canc_{f}")
        canceladas_por_fecha[f] = es_cancelada
        
        if not es_cancelada:
            visita_sc_pasada = fecha_efectiva > FECHA_VISITA_SC
            es_septiembre_o_mas = fecha_efectiva.month >= 9
            
            disp_av = hermanos_av.copy()
            if visita_sc_pasada and "Geremy Fernández" in disp_av:
                disp_av.remove("Geremy Fernández")
                
            disp_mics = disp_av + hermanos_solo_mics
            if es_septiembre_o_mas:
                disp_mics.append("Iván Chavarría")

            disp_acom = ancianos_y_ministeriales.copy()
            if visita_sc_pasada:
                if "Geremy Fernández" in disp_acom:
                    disp_acom.remove("Geremy Fernández")
                if "Roger Loaiza" in disp_acom:
                    disp_acom.remove("Roger Loaiza")

            opciones_totales = sorted(list(set(disp_av + disp_mics + disp_acom)))
            
            ocupados_por_fecha[f] = st.multiselect(
                f"Hermanos NO disponibles el {dia_real} {fecha_efectiva.strftime('%d/%m/%Y')}:",
                options=opciones_totales,
                key=f.strftime("%Y-%m-%d")
            )
        else:
            ocupados_por_fecha[f] = []
        
        st.divider()

def seleccionar_equilibrado(lista_candidatos, contador_usos, ultimos_asignados, cantidad=1):
    if not lista_candidatos:
        return []
    candidatos_validos = [h for h in lista_candidatos if not (h == "Roger Loaiza" and contador_usos[h] >= 1)]
    if not candidatos_validos:
        return []
    def peso_prioridad(h):
        penalizacion_reciente = 2.0 if h in ultimos_asignados else 0.0
        return (contador_usos[h] + penalizacion_reciente, random.random())
    return sorted(candidatos_validos, key=peso_prioridad)[:cantidad]

# --- 4. Generación ---
if st.button("🚀 Generar Programa Completo"):
    filas_programa = []
    error_detectado = False
    contador_usos = Counter()
    ultimos_asignados = hermanos_recientes.copy()
    
    for f in fechas_reunion:
        dia_nombre, fecha_efectiva = dias_ajustados_por_fecha[f]
        fecha_txt = fecha_efectiva.strftime("%d/%m/%Y")
        
        if canceladas_por_fecha[f]:
            filas_programa.append({
                "Fecha": fecha_txt, "Día": dia_nombre,
                "Audio": "--- NO HAY REUNIÓN ---", "Video": "--- NO HAY REUNIÓN ---",
                "Micrófono": "--- NO HAY REUNIÓN ---", "Acomodador": "--- NO HAY REUNIÓN ---"
            })
            ultimos_asignados = []
            continue

        visita_sc_pasada = fecha_efectiva > FECHA_VISITA_SC
        es_septiembre_o_mas = fecha_efectiva.month >= 9
        
        d_av = [h for h in hermanos_av if not (visita_sc_pasada and h == "Geremy Fernández")]
        d_mics = d_av + hermanos_solo_mics + (["Iván Chavarría"] if es_septiembre_o_mas else [])
        d_acom = [h for h in ancianos_y_ministeriales if not (visita_sc_pasada and h in ["Geremy Fernández", "Roger Loaiza"])]
            
        ocupados_hoy = ocupados_por_fecha[f]
        libres_av = [h for h in d_av if h not in ocupados_hoy]
        libres_mics = [h for h in d_mics if h not in ocupados_hoy]
        libres_acom = [h for h in d_acom if h not in ocupados_hoy]
        
        if len(libres_av) >= 2 and len(libres_mics) >= 1 and len(libres_acom) >= 1:
            equipo_av = seleccionar_equilibrado(libres_av, contador_usos, ultimos_asignados, cantidad=2)
            for h in equipo_av: contador_usos[h] += 1
            
            libres_mics_rest = [h for h in libres_mics if h not in equipo_av]
            equipo_mics = seleccionar_equilibrado(libres_mics_rest, contador_usos, ultimos_asignados, cantidad=1)
            for h in equipo_mics: contador_usos[h] += 1
            
            libres_acom_rest = [h for h in libres_acom if h not in equipo_av and h not in equipo_mics]
            equipo_acom = seleccionar_equilibrado(libres_acom_rest, contador_usos, ultimos_asignados, cantidad=1) or ["Revisar manual"]
            if equipo_acom[0] != "Revisar manual": contador_usos[equipo_acom[0]] += 1
            
            ultimos_asignados = equipo_av + equipo_mics + equipo_acom
            filas_programa.append({
                "Fecha": fecha_txt, "Día": dia_nombre,
                "Audio": equipo_av[0], "Video": equipo_av[1],
                "Micrófono": equipo_mics[0], "Acomodador": equipo_acom[0]
            })
        else:
            st.error(f"Faltan hermanos disponibles para el {fecha_txt}.")
            error_detectado = True

    if not error_detectado:
        st.session_state["programa_guardado"] = pd.DataFrame(filas_programa)
        st.session_state["contador_usos"] = contador_usos
        st.success("¡Programa generado con éxito!")

# --- 5. Exportaciones y Descargas ---
if st.session_state["programa_guardado"] is not None:
    st.divider()
    df_editable = st.data_editor(st.session_state["programa_guardado"], use_container_width=True)

    # Funciones de Generación
    def generar_imagen(df, titulo_rango, logo_bytes=None, es_jpg=False):
        img_w, row_h, head_h, th_h = 1000, 42, 130, 45
        tot_h = head_h + th_h + len(df) * row_h + 60
        img = Image.new("RGB", (img_w, tot_h), color=(255, 255, 255))
        draw = ImageDraw.Draw(img)

        try:
            f_title = ImageFont.truetype("DejaVuSans-Bold.ttf", 24)
            f_sub = ImageFont.truetype("DejaVuSans.ttf", 15)
            f_th = ImageFont.truetype("DejaVuSans-Bold.ttf", 14)
            f_td = ImageFont.truetype("DejaVuSans.ttf", 13)
        except:
            f_title = f_sub = f_th = f_td = ImageFont.load_default()

        draw.rectangle([(0, 0), (img_w, head_h)], fill=(31, 78, 120))
        if logo_bytes:
            try:
                logo_img = Image.open(io.BytesIO(logo_bytes)).convert("RGBA")
                logo_img.thumbnail((90, 90))
                img.paste(logo_img, (35, 20), logo_img)
            except Exception: pass

        draw.text((img_w // 2, 45), "PROGRAMA DE AUDIO, VIDEO Y SALAS", fill=(255, 255, 255), font=f_title, anchor="mm")
        draw.text((img_w // 2, 85), f"Congregación Gallito • {titulo_rango}", fill=(210, 230, 250), font=f_sub, anchor="mm")

        col_x = [30, 150, 270, 450, 630, 810, 970]
        draw.rectangle([(col_x[0], head_h + 10), (col_x[-1], head_h + 10 + th_h)], fill=(44, 62, 80))

        for i, h_text in enumerate(["Fecha", "Día", "Audio", "Video", "Micrófono", "Acomodador"]):
            draw.text(((col_x[i] + col_x[i+1]) // 2, head_h + 10 + (th_h // 2)), h_text, fill=(255, 255, 255), font=f_th, anchor="mm")

        y_curr = head_h + 10 + th_h
        for idx, (_, r) in enumerate(df.iterrows()):
            bg = (248, 250, 252) if idx % 2 == 1 else (255, 255, 255)
            draw.rectangle([(col_x[0], y_curr), (col_x[-1], y_curr + row_h)], fill=bg, outline=(226, 232, 240))
            vals = [str(r["Fecha"]), str(r["Día"]), str(r["Audio"]), str(r["Video"]), str(r["Micrófono"]), str(r["Acomodador"])]
            for c_i, val in enumerate(vals):
                cx = (col_x[c_i] + col_x[c_i+1]) // 2 if c_i < 2 else col_x[c_i] + 12
                anchor = "mm" if c_i < 2 else "lm"
                draw.text((cx, y_curr + (row_h // 2)), val, fill=(44, 62, 80), font=f_td, anchor=anchor)
            y_curr += row_h

        buf = io.BytesIO()
        img.save(buf, format="JPEG" if es_jpg else "PNG", quality=95 if es_jpg else None)
        return buf.getvalue()

    def generar_excel(df, titulo_rango):
        buf = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Programa"
        header_fill = openpyxl.styles.PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = openpyxl.styles.Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        ws.merge_cells("A1:F1")
        ws["A1"] = f"PROGRAMA DE AUDIO, VIDEO Y SALAS - {titulo_rango}"
        ws["A1"].font = openpyxl.styles.Font(size=14, bold=True, color="1F4E78")
        ws["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")

        cols = ["Fecha", "Día", "Audio", "Video", "Micrófono", "Acomodador"]
        for c_i, h in enumerate(cols, 1):
            cell = ws.cell(row=3, column=c_i, value=h)
            cell.fill, cell.font = header_fill, header_font
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        for r_i, (_, r) in enumerate(df.iterrows(), 4):
            for c_i, h in enumerate(cols, 1):
                cell = ws.cell(row=r_i, column=c_i, value=str(r[h]))
                cell.alignment = openpyxl.styles.Alignment(horizontal="center" if c_i <= 2 else "left")

        for col in ws.columns:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = 20

        wb.save(buf)
        return buf.getvalue()

    def generar_pdf(df, titulo_rango, logo_bytes=None):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=25, leftMargin=25, topMargin=25, bottomMargin=25)
        elements = []
        styles = getSampleStyleSheet()

        if logo_bytes:
            try:
                rl_img = RLImage(io.BytesIO(logo_bytes), width=50, height=50)
                rl_img.hAlign = 'CENTER'
                elements.append(rl_img)
            except Exception: pass

        elements.append(Paragraph(f"<b>PROGRAMA DE AUDIO, VIDEO Y SALAS</b><br/><font size=10 color='#555555'>{titulo_rango}</font>", styles['Heading2']))
        elements.append(Spacer(1, 10))

        datos = [["Fecha", "Día", "Audio", "Video", "Micrófono", "Acomodador"]]
        for _, row in df.iterrows():
            datos.append([str(row['Fecha']), str(row['Día']), str(row['Audio']), str(row['Video']), str(row['Micrófono']), str(row['Acomodador'])])

        t = Table(datos, colWidths=[65, 65, 110, 110, 105, 105])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
            ('FONTSIZE', (0,0), (-1,-1), 8.5),
        ]))
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # --- TRES BOTONES DE DESCARGA ---
    st.subheader("📥 DESCARGAR PROGRAMA COMPLETO")
    c1, c2, c3 = st.columns(3)
    logo_data = logo_subido.getvalue() if logo_subido else None

    if EXCEL_DISPONIBLE:
        c1.download_button("📊 Excel (.xlsx)", generar_excel(df_editable, texto_rango), f"Programa_{mes_sel}.xlsx", use_container_width=True)

    if PDF_DISPONIBLE:
        c2.download_button("📄 PDF (Imprimir)", generar_pdf(df_editable, texto_rango, logo_data), f"Programa_{mes_sel}.pdf", use_container_width=True)

    c3.download_button("🖼️ Imagen (PNG)", generar_imagen(df_editable, texto_rango, logo_data, es_jpg=False), f"Programa_{mes_sel}.png", use_container_width=True)
